# src/core/file_handlers/excel_handler.py
"""
Enhanced Excel handler with smart OCR support for embedded images.
"""

import os
import logging
from typing import List, Tuple
from pathlib import Path
import tempfile
from zipfile import ZipFile
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import xlrd
from PIL import Image
import io
import asyncio

from src.core.file_handlers.base_handler import FileHandler

# Import hybrid OCR system
try:
    from ..services.ocr_service import get_ocr_service
    from ..utils.config_loader import ConfigLoader
    HYBRID_OCR_AVAILABLE = True
except ImportError:
    HYBRID_OCR_AVAILABLE = False

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.DEBUG)

class ExcelHandler(FileHandler):
    """
    Enhanced Excel handler with smart OCR for embedded images and charts.
    """

    def __init__(self, model_manager=None, use_smart_ocr=True):
        """Initialize the Excel handler with smart OCR capabilities."""
        self.model_manager = model_manager
        self.temp_dir = tempfile.TemporaryDirectory()
        
        # Hybrid OCR integration
        self.use_smart_ocr = use_smart_ocr and HYBRID_OCR_AVAILABLE
        if self.use_smart_ocr:
            try:
                self.ocr_service = get_ocr_service()
                self.ocr_config = ConfigLoader.load_ocr_config()
                config_status = self.ocr_service.get_config_status()
                self.smart_ocr_available = config_status["tesseract_available"] or config_status["easyocr_available"]
                logger.info(f"Excel Handler smart OCR enabled: {self.smart_ocr_available}")
            except Exception as e:
                logger.warning(f"Smart OCR initialization failed: {e}, falling back to legacy mode")
                self.use_smart_ocr = False
                self.smart_ocr_available = False
        
    async def extract_text(self, file_path: str, sheet_name: str = None) -> str:
        """Extract text from an Excel file with smart image processing."""
        if not os.path.exists(file_path):
            logger.error(f"File not found: {file_path}")
            raise FileNotFoundError(f"File not found: {file_path}")

        ext = Path(file_path).suffix.lower()
        if ext not in ['.xls', '.xlsx']:
            logger.error(f"Unsupported file format: {file_path}")
            raise ValueError("Unsupported file format. Only .xls and .xlsx are supported.")

        try:
            # Phase 1: Extract text from sheets
            sheet_text = await self._extract_text_from_sheets(file_path, sheet_name)
            total_text_length = len(sheet_text)
            
            # Phase 2: Determine OCR strategy for embedded images
            should_process_images, ocr_mode = self._determine_smart_ocr_strategy(
                total_text_length, 1, "xlsx"
            )
            
            text_parts = [sheet_text] if sheet_text else []
            
            if should_process_images and ext == '.xlsx':  # Only XLSX supports embedded images easily
                if self.use_smart_ocr and self.smart_ocr_available:
                    logger.info(f"Excel has {total_text_length} chars, using smart OCR mode: {ocr_mode}")
                    image_text = await self._extract_images_smart(file_path, ocr_mode)
                else:
                    # Excel rarely has meaningful images, but try basic extraction
                    image_text = await self._extract_images_basic(file_path)
                
                if image_text:
                    text_parts.append(f"=== EMBEDDED IMAGES ===\n{image_text}")
            else:
                logger.info(f"Excel has sufficient text ({total_text_length} chars) or is XLS format, skipping image OCR")
            
            logger.debug(f"Successfully extracted text from {file_path}")
            return "\n\n".join(text_parts) if text_parts else ""
            
        except Exception as e:
            logger.error(f"Error processing Excel file {file_path}: {e}")
            return ""

    async def _extract_images_smart(self, file_path: str, ocr_mode: str) -> str:
        """Extract images from Excel with smart OCR strategy."""
        try:
            # Extract images from the Excel file
            images = await self._extract_excel_images(file_path)
            
            if not images:
                return ""
            
            # Excel images are usually charts, diagrams, or screenshots
            context = "table"  # Excel context is usually tabular data
            
            # Use smart OCR service
            engine = "tesseract"  # Tesseract is usually better for structured Excel content
            extracted_texts = await self.ocr_service.extract_text_from_images(
                images, ocr_mode, context, engine
            )
            
            # Format the results
            image_texts = []
            for i, text in enumerate(extracted_texts):
                if text:
                    image_texts.append(f"Chart/Image {i + 1}: {text}")
            
            logger.debug(f"Smart OCR processed {len(image_texts)} images from Excel")
            return "\n".join(image_texts)
            
        except Exception as e:
            logger.error(f"Smart image extraction from Excel failed: {e}")
            return ""

    async def _extract_excel_images(self, file_path: str) -> List[Image.Image]:
        """Extract images from Excel file."""
        images = []
        
        try:
            # Excel images are embedded in the file structure
            # We need to extract them from the internal structure
            
            def sync_extract():
                workbook = openpyxl.load_workbook(file_path)
                extracted_images = []
                
                try:
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        
                        # Check for images in the sheet
                        if hasattr(sheet, '_images') and sheet._images:
                            for img in sheet._images:
                                try:
                                    # Extract image data
                                    if hasattr(img, 'ref') and hasattr(img, '_data'):
                                        image_data = img._data()
                                        if image_data:
                                            pil_image = Image.open(io.BytesIO(image_data))
                                            extracted_images.append(pil_image)
                                except Exception as e:
                                    logger.debug(f"Failed to extract image from sheet {sheet_name}: {e}")
                                    continue
                
                except Exception as e:
                    logger.debug(f"Error accessing sheet images: {e}")
                
                finally:
                    workbook.close()
                
                return extracted_images
            
            # Run in thread pool
            images = await asyncio.to_thread(sync_extract)
            
        except Exception as e:
            logger.debug(f"Error extracting images from Excel: {e}")
        
        return images

    async def _extract_images_basic(self, file_path: str) -> str:
        """Basic image extraction method (placeholder)."""
        # Excel image extraction is complex and rarely contains meaningful text
        # This is a placeholder for basic image extraction
        return ""

    def _determine_smart_ocr_strategy(self, text_length: int, sheet_count: int, file_type: str) -> tuple[bool, str]:
        """Determine if smart OCR should be used for Excel."""
        if not self.use_smart_ocr or not self.smart_ocr_available:
            return False, "skip"  # Excel rarely needs OCR
            
        # Excel files rarely have meaningful images with text
        # Only process if very little text is present (indicating chart-heavy file)
        if text_length < 200:
            return True, "selective"  # Might be chart-heavy
        elif text_length < 1000:
            return True, "minimal"   # Some charts possible
        else:
            return False, "skip"     # Text-rich spreadsheet

    async def _extract_text_from_sheets(self, file_path: str, sheet_name: str = None) -> str:
        """Extract text from Excel sheets."""
        ext = Path(file_path).suffix.lower()
        if ext == '.xlsx':
            return await self._extract_from_xlsx(file_path, sheet_name)
        elif ext == '.xls':
            return await self._extract_from_xls(file_path, sheet_name)
        return ""

    async def _extract_from_xlsx(self, file_path: str, sheet_name: str = None) -> str:
        """Extract text from XLSX file."""
        import openpyxl
        
        # Use asyncio to run this in a thread pool since openpyxl is synchronous
        def _read_xlsx():
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            try:
                text = []
                sheets = [sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.sheetnames
                
                for sheet in sheets:
                    worksheet = workbook[sheet]
                    text.append(f"=== Sheet: {sheet} ===")
                    
                    for row in worksheet.iter_rows():
                        row_text = []
                        for cell in row:
                            if cell.value is not None:
                                row_text.append(str(cell.value))
                        
                        if row_text:
                            text.append('\t'.join(row_text))
                
                return '\n'.join(text)
            finally:
                workbook.close()
                
        # Run synchronous code in a thread pool
        return await asyncio.to_thread(_read_xlsx)

    async def _extract_from_xls(self, file_path: str, sheet_name: str = None) -> str:
        """Extract text from XLS file."""
        import xlrd
        
        # Use asyncio to run this in a thread pool since xlrd is synchronous
        def _read_xls():
            workbook = xlrd.open_workbook(file_path)
            text = []
            
            sheets = [sheet_name] if sheet_name and sheet_name in workbook.sheet_names() else workbook.sheet_names()
            
            for sheet in sheets:
                worksheet = workbook.sheet_by_name(sheet)
                text.append(f"=== Sheet: {sheet} ===")
                
                for row_idx in range(worksheet.nrows):
                    row_text = []
                    for col_idx in range(worksheet.ncols):
                        value = worksheet.cell_value(row_idx, col_idx)
                        if value:
                            row_text.append(str(value))
                    
                    if row_text:
                        text.append('\t'.join(row_text))
            
            return '\n'.join(text)
            
        # Run synchronous code in a thread pool
        return await asyncio.to_thread(_read_xls)

    async def extract_tables(self, file_path: str, sheet_name: str = None) -> List[List[List[str]]]:
        """Extract tables from an Excel file (each sheet as a table)."""
        if not os.path.exists(file_path):
            logger.error(f"File not found: {file_path}")
            raise FileNotFoundError(f"File not found: {file_path}")

        ext = Path(file_path).suffix.lower()
        if ext not in ['.xls', '.xlsx']:
            logger.error(f"Unsupported file format: {file_path}")
            raise ValueError("Unsupported file format. Only .xls and .xlsx are supported.")

        try:
            if ext == '.xlsx':
                return await self._extract_tables_from_xlsx(file_path, sheet_name)
            elif ext == '.xls':
                return await self._extract_tables_from_xls(file_path, sheet_name)
            return []
        except Exception as e:
            logger.error(f"Error extracting tables from {file_path}: {e}")
            return []

    async def _extract_tables_from_xlsx(self, file_path: str, sheet_name: str = None) -> List[List[List[str]]]:
        """Extract tables from XLSX file."""
        import openpyxl
        
        # Use asyncio to run this in a thread pool
        def _read_tables_xlsx():
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            try:
                tables = []
                sheets = [sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.sheetnames
                
                for sheet in sheets:
                    worksheet = workbook[sheet]
                    table_data = []
                    
                    for row in worksheet.iter_rows():
                        row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                        if any(row_data):  # Skip empty rows
                            table_data.append(row_data)
                    
                    if table_data:  # Skip empty tables
                        tables.append(table_data)
                
                return tables
            finally:
                workbook.close()
                
        # Run synchronous code in a thread pool
        return await asyncio.to_thread(_read_tables_xlsx)

    async def _extract_tables_from_xls(self, file_path: str, sheet_name: str = None) -> List[List[List[str]]]:
        """Extract tables from XLS file."""
        import xlrd
        
        # Use asyncio to run this in a thread pool
        def _read_tables_xls():
            workbook = xlrd.open_workbook(file_path)
            tables = []
            
            sheets = [sheet_name] if sheet_name and sheet_name in workbook.sheet_names() else workbook.sheet_names()
            
            for sheet in sheets:
                worksheet = workbook.sheet_by_name(sheet)
                table_data = []
                
                for row_idx in range(worksheet.nrows):
                    row_data = [str(worksheet.cell_value(row_idx, col_idx)) if worksheet.cell_value(row_idx, col_idx) else "" 
                               for col_idx in range(worksheet.ncols)]
                    if any(row_data):  # Skip empty rows
                        table_data.append(row_data)
                
                if table_data:  # Skip empty tables
                    tables.append(table_data)
            
            return tables
            
        # Run synchronous code in a thread pool
        return await asyncio.to_thread(_read_tables_xls)

    def get_handler_status(self) -> dict:
        """Get current handler status for debugging."""
        status = {
            "handler": "Enhanced Excel Handler",
            "smart_ocr_enabled": self.use_smart_ocr,
            "smart_ocr_available": getattr(self, 'smart_ocr_available', False),
            "supported_formats": [".xls", ".xlsx"]
        }
        
        if self.use_smart_ocr and hasattr(self, 'ocr_service'):
            status["ocr_service_status"] = self.ocr_service.get_config_status()
        
        return status

    def __del__(self):
        """Clean up temporary directory when the handler is destroyed."""
        if hasattr(self, 'temp_dir') and self.temp_dir:
            self.temp_dir.cleanup()